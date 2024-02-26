""" This module contains the class for reading """

from datetime import datetime
import openpyxl
from .constants import INDICE_TRADUCCION, INDICE_TRADUCCION_CONTACT

from auxiliares.models import ContactType, EmailType, SocialType, WebType, Country, Type
from .models import (Clientes,
                     ClientesAddress,
                     ClientesContact,
                     ClientesEmail,
                     ClientesSocial,
                     ClientesWeb,
                     MailCorp,
                     Account)



class UtilExcelFile():
    '''
    Class for reading and manipulating Excel files.

    Methods:
    - __init__: Initializes the class with an empty workbook and active worksheet.
    - open_file: Opens an Excel file and sets the workbook and active worksheet.
    - clean_name: Cleans the name of a column by removing spaces, dashes, periods, 
        commas, parentheses, and slashes.
    - get_structure: Returns a dictionary with the structure of the Excel file and a 
        dictionary with the indices.
    - get_data: Returns a dictionary with the data of the Excel file and a dictionary 
        with the indices.
    - print_datos: Prints the data of the Excel file to the console and creates and 
        saves instances of various models in a Django project based on the data.
    - add_sheet: Adds a sheet to the Excel file.
    - add_row: Adds a row to the active worksheet.
    - save: Saves the Excel file.

    Fields:
    - file_name: The name of the Excel file.
    - wb: The openpyxl workbook object.
    - ws: The openpyxl worksheet object.
    '''
    file_name = None
    wb = None
    ws = None

    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def open_file(self, file_name):
        ''' Abre un archivo excel '''
        try:
            self.file_name = file_name
            self.wb = openpyxl.load_workbook(file_name)
            self.ws = self.wb.active
        except FileNotFoundError:
            print("El archivo no existe")
        except Exception as e:
            print("Error al abrir el archivo: " + str(e))


    def clean_name(self, name):
        ''' Limpia el nombre de los campos '''
        name = str(name).lower()
        name = name.replace(" ", "_")
        name = name.replace("-", "_")
        name = name.replace(".", "")
        name = name.replace(",", "")
        name = name.replace("(", "")
        name = name.replace(")", "")
        name = name.replace("/", "")
        return name

    def traducir_claves_dict(self, clave, diccionario):
        '''
        Diccionario de traducción
        '''
        if clave in diccionario:
            return diccionario[clave]
        else:
            return clave

    def convertir_a_datetime(self, fecha_str):
        """ Convierte una fecha en string a datetime """
        formatos = ["%Y-%m-%d",
                    "%d/%m/%Y",
                    "%m-%d-%Y",
                    "%m/%d/%y %H:%M",
                    "%m/%d/%y %H:%M:%S %p",
                    "%m/%d/%Y %H:%M",
                    "%m/%d/%Y %H:%M:%S %p"]
        for formato in formatos:
            try:
                if fecha_str is None:
                    return None
                elif isinstance(fecha_str, datetime):
                    return fecha_str
                fecha_datetime = datetime.strptime(fecha_str, formato)
                return fecha_datetime
            except ValueError:
                pass

        raise ValueError(f"No se pudo encontrar un formato válido para la fecha: {fecha_str}")



    def get_structure(self):
        ''' Devuelve un diccionario con la estructura del excel y un diccionario con los indices '''
        response = {}
        indice = {}
        for row in self.ws.iter_rows(min_row=1, max_row=1):
            i = 0
            for cell in row:
                indice[i] = cell.value

                if cell.value:
                    indio = self.clean_name(cell.value)
                    indio = self.traducir_claves_dict(indio, INDICE_TRADUCCION)
                    # TODO: Este if es una excepción, hay que ver como solucionarlo
                    if i == 58:
                        indio = indio+"_dire"
                else:
                    indio = i

                response[indio] = []
                i += 1

        return response, indice

    def get_data(self) -> tuple:
        ''' Devuelve un diccionario con los datos del excel y un diccionario con los indices '''
        try:
            estructura, indice = self.get_structure()

            for row in self.ws.iter_rows(min_row=2):
                i=0
                # for i in range(len(row)):
                for campo in row:
                    if indice[i]:
                        indio = self.clean_name(indice[i])
                        indio = self.traducir_claves_dict(indio, INDICE_TRADUCCION)
                        indio = indio + "_dire" if i == 58 else indio
                    else:
                        indio = i
                    dato = campo.value
                    estructura[indio].append(dato)
                    i += 1

        except KeyError as e:
            print("Error de key error: " + str(e))
            raise ValueError(str(e))
        except Exception as e:
            print("Error al obtener los datos: " + str(e))
            raise ValueError(str(e))


        return estructura, indice

    def print_datos(self):
        ''' Imprime los datos del excel '''
        data, indice = self.get_data()
        for i in range(len(data[list(data.keys())[0]])):

            if data['id'][i] is None:
                print("No hay id en posicion: " + str(i))
                continue
            if not Clientes.objects.filter(cliente_id=data['id'][i]).exists():
                cliente = Clientes()
                cliente.cliente_id = data['id'][i]
                action = "creado"
            else:
                try:
                    cliente = Clientes.objects.get(cliente_id=data['id'][i])
                    action = "actualizado"
                except Clientes.MultipleObjectsReturned as exc:
                    cliente = Clientes.objects.filter(cliente_id=data['id'][i]).first()
                    mensaje = f"Se encontraron multiples clientes con el id:{data['id'][i]}"
                    raise ValueError(mensaje) from exc

            cliente.status = self.get_status(data, i)
            cliente.lead_name = self.get_lead_name(data, i)
            cliente.salutation = self.get_salutation(data, i)
            cliente.first_name = self.get_first_name(data, i)
            cliente.middle_name = self.get_middle_name(data, i)
            cliente.last_name = self.get_last_name(data, i)
            cliente.date_of_birth = self.get_date_of_birth(data, i)
            cliente.created = self.get_created(data, i)
            cliente.source = self.get_source(data, i)
            cliente.responsible = self.get_responsible(data, i)
            cliente.status_information = self.get_status_information(data, i)
            cliente.source_information = self.get_source_information(data, i)
            cliente.created_by = self.get_created_by(data, i)
            cliente.modified = self.get_modified(data, i)
            cliente.modified_by = self.get_modified_by(data, i)
            cliente.company_name = self.get_company_name(data, i)
            cliente.position = self.get_position(data, i)
            cliente.comment = self.get_comment(data, i)
            cliente.total = self.get_total(data, i)
            cliente.currency = self.get_currency(data, i)
            cliente.product = self.get_product(data, i)
            cliente.price = self.get_price(data, i)
            cliente.quantity = self.get_quantity(data, i)
            cliente.created_by_crm_form = self.get_created_by_crm_form(data, i)
            cliente.repeat_lead = self.get_repeat_lead(data, i)
            cliente.client = self.get_client(data, i)
            cliente.customer_journey = self.get_customer_journey(data, i)
            cliente.type = self.get_type(data, i)
            cliente.country = self.get_country(data, i)
            cliente.account = self.get_account(data, i)
            cliente.addl_type_details_other = self.get_addl_type_details_other(data, i)
            cliente.industry_sub_type = self.get_industry_sub_type(data, i)
            cliente.last_updated_on = self.get_last_updated_on(data, i)

            cliente.save()

            direccion = ClientesAddress()
            direccion.cliente = cliente
            direccion.address = self.get_address(data, i)
            direccion.street_house_no = self.get_street_house_no(data, i)
            direccion.apartment_office_room_floor = self.get_apartment_office_room_floor(data, i)
            direccion.city = self.get_city(data, i)
            direccion.district = self.get_district(data, i)
            direccion.region_area = self.get_regionarea(data, i)
            direccion.postal_code = self.get_zippostal_code(data, i)
            direccion.country = self.get_country_dire(data, i)

            direccion.save()
            e = 0
            for clave, item in data.items():
                if item[i]:
                    nombre = self.traducir_claves_dict(indice[e], INDICE_TRADUCCION_CONTACT)
                    if nombre and ContactType.objects.filter(name=nombre):
                        if not ClientesContact.objects.filter(cliente=cliente,
                                                              type=ContactType.objects.get(name=nombre)).exists():
                            cliente.add_contact(
                                ContactType.objects.get(name=nombre), item[i])
                        else:
                            contact = ClientesContact.objects.get(
                                cliente=cliente, type=ContactType.objects.get(name=nombre))
                            contact.data = item[i]
                            contact.save()

                    elif nombre and WebType.objects.filter(name=nombre):
                        if not ClientesWeb.objects.filter(cliente=cliente, type=WebType.objects.get(name=nombre)).exists():
                            cliente.add_web(
                                WebType.objects.get(name=nombre), item[i])
                        else:
                            contact = ClientesWeb.objects.get(
                                cliente=cliente, type=WebType.objects.get(name=nombre))
                            contact.data = item[i]
                            contact.save()

                    elif nombre and EmailType.objects.filter(name=nombre):
                        if not ClientesEmail.objects.filter(cliente=cliente, type=EmailType.objects.get(name=nombre)).exists():
                            cliente.add_email(
                                EmailType.objects.get(name=nombre), item[i])
                        else:
                            contact = ClientesEmail.objects.get(
                                cliente=cliente, type=EmailType.objects.get(name=nombre))
                            contact.data = item[i]
                            contact.save()

                    elif nombre and SocialType.objects.filter(name=nombre):
                        if not ClientesSocial.objects.filter(cliente=cliente, type=SocialType.objects.get(name=nombre)).exists():
                            cliente.add_social(
                                SocialType.objects.get(name=nombre), item[i])
                        else:
                            contact = ClientesSocial.objects.get(
                                cliente=cliente, type=SocialType.objects.get(name=nombre))
                            contact.data = item[i]
                            contact.save()

                e += 1

            # cliente.save()
            print("Se ha "+action+" el cliente: ",
                  cliente.first_name, cliente.last_name)

    def add_sheet(self, sheet_name):
        ''' Agrega una hoja al archivo'''
        self.ws = self.wb.create_sheet(sheet_name)

    def add_row(self, row):
        ''' Agrega una fila al archivo '''
        self.ws.append(row)

    def save(self):
        ''' Guarda el archivo '''
        self.wb.save(self.file_name)

    def get_status(self, data: list, indice: int)-> str:
        '''
        This method is used to get the status from the data
        '''
        if 'status' in data:
            return data['status'][indice]
        elif 'estatus' in data:
            return data['estatus'][indice]
        elif 'stage' in data:
            return data['stage'][indice]
        elif 'etapa' in data:
            return data['etapa'][indice]
        else:
            return None

    def get_lead_name(self, data: list, indice: int)-> str:
        '''
        This method is used to get the lead name from the data
        '''
        if 'lead_name' in data:
            return data['lead_name'][indice]
        elif 'título_de_prospecto' in data:
            return data['título_de_prospecto'][indice]
        else:
            return None

    def get_salutation(self, data: list, indice: int)-> str:
        '''
        This method is used to get the salutation from the data
        '''
        if 'salutation' in data:
            return data['salutation'][indice]
        elif 'saludo' in data:
            return data['saludo'][indice]
        else:
            return None

    def get_first_name(self, data: list, indice: int)-> str:
        '''
        This method is used to get the first name from the data
        '''
        if 'first_name' in data:
            return data['first_name'][indice]
        elif 'nombre' in data:
            return data['nombre'][indice]
        else:
            return None

    def get_middle_name(self, data: list, indice: int)-> str:
        '''
        This method is used to get the middle name from the data
        '''
        if 'middle_name' in data:
            return data['middle_name'][indice]
        elif 'segundo_nombre' in data:
            return data['segundo_nombre'][indice]
        else:
            return None

    def get_last_name(self, data: list, indice: int)-> str:
        '''
        This method is used to get the last name from the data
        '''
        if 'last_name' in data:
            return data['last_name'][indice]
        elif 'apellido' in data:
            return data['apellido'][indice]
        else:
            return None

    def get_date_of_birth(self, data: list, indice: int)-> str:
        '''
        This method is used to get the date of birth from the data
        '''
        if 'date_of_birth' in data:
            return data['date_of_birth'][indice]
        elif 'fecha_de_nacimiento' in data:
            return data['fecha_de_nacimiento'][indice]
        else:
            return None

    def get_created(self, data: list, indice: int)-> str:
        '''
        This method is used to get the created from the data
        '''
        if 'created' in data:
            return self.convertir_a_datetime(data['created'][indice])
        elif 'creado' in data:
            return self.convertir_a_datetime(data['creado'][indice])
        else:
            return None

    def get_source(self, data: list, indice: int)-> str:
        '''
        This method is used to get the source from the data
        '''
        if 'source' in data:
            return data['source'][indice]
        elif 'origen' in data:
            return data['origen'][indice]
        else:
            return None

    def get_responsible(self, data: list, indice: int)-> str:
        '''
        This method is used to get the responsible from the data
        '''
        try:
            if 'responsible' in data:
                responsible = data['responsible'][indice]
            elif 'responsable' in data:
                responsible = data['responsable'][indice]
            else:
                return None

            return MailCorp.objects.get(name=responsible)
        except MailCorp.DoesNotExist as exc:
            raise ValueError('The responsible does not exist: ' + responsible) from exc

    def get_status_information(self, data: list, indice: int)-> str:
        '''
        This method is used to get the status information from the data
        '''
        if 'status_information' in data:
            return data['status_information'][indice]
        elif 'información_de_estado' in data:
            return data['información_de_estado'][indice]
        else:
            return None

    def get_source_information(self, data: list, indice: int)-> str:
        '''
        This method is used to get the source information from the data
        '''
        try:
            if 'source_information' in data:
                return data['source_information'][indice]
            elif 'información_de_origen' in data:
                return data['información_de_origen'][indice]
            else:
                return None
        except MailCorp.DoesNotExist as exc:
            raise ValueError('The source_information does not exist: ' +
                             data['información_de_origen'][indice]) from exc
        except Exception as exc:
            raise ValueError('Error en source_information: ' +
                             data['información_de_origen'][indice]) from exc

    def get_created_by(self, data: list, indice: int)-> str:
        '''
        This method is used to get the created by from the data
        '''
        try:
            if 'created_by' in data:
                created_by = data['created_by'][indice]
            elif 'creado_por' in data:
                created_by = data['creado_por'][indice]
            else:
                return None

            return MailCorp.objects.get(name=created_by)
        except MailCorp.DoesNotExist as exc:
            raise ValueError('The created_by does not exist: ' + created_by) from exc

    def get_modified(self, data: list, indice: int)-> str:
        '''
        This method is used to get the modified from the data
        '''
        if 'modified' in data:
            return self.convertir_a_datetime(data['modified'][indice])
        elif 'modificado' in data:
            return self.convertir_a_datetime(data['modificado'][indice])
        else:
            return None

    def get_modified_by(self, data: list, indice: int)-> str:
        '''
        This method is used to get the modified by from the data
        '''
        try:
            if 'modified_by' in data:
                modified_by = data['modified_by'][indice]
            elif 'modificado_por' in data:
                modified_by = data['modificado_por'][indice]
            else:
                return None

            return MailCorp.objects.get(name=modified_by)
        except MailCorp.DoesNotExist as exc:
            raise ValueError('The modified_by does not exist: ' + modified_by) from exc

    def get_company_name(self, data: list, indice: int)-> str:
        '''
        This method is used to get the company name from the data
        '''
        if 'company_name' in data:
            return data['company_name'][indice]
        elif 'nombre_de_la_compañía' in data:
            return data['nombre_de_la_compañía'][indice]
        else:
            return None

    def get_position(self, data: list, indice: int)-> str:
        '''
        This method is used to get the position from the data
        '''
        if 'position' in data:
            return data['position'][indice]
        elif 'cargo' in data:
            return data['cargo'][indice]
        else:
            return None

    def get_comment(self, data: list, indice: int)-> str:
        '''
        This method is used to get the comment from the data
        '''
        if 'comment' in data:
            return data['comment'][indice]
        elif 'comentario' in data:
            return data['comentario'][indice]
        else:
            return None

    def get_total(self, data: list, indice: int)-> str:
        '''
        This method is used to get the total from the data
        '''
        if 'total' in data:
            return data['total'][indice]
        else:
            return None

    def get_currency(self, data: list, indice: int)-> str:
        '''
        This method is used to get the currency from the data
        '''
        if 'currency' in data:
            return data['currency'][indice]
        elif 'moneda' in data:
            return data['moneda'][indice]
        else:
            return None

    def get_product(self, data: list, indice: int)-> str:
        '''
        This method is used to get the product from the data
        '''
        if 'product' in data:
            return data['product'][indice]
        elif 'producto' in data:
            return data['producto'][indice]
        else:
            return None

    def get_price(self, data: list, indice: int)-> str:
        '''
        This method is used to get the price from the data
        '''
        if 'price' in data:
            return data['price'][indice]
        elif 'precio' in data:
            return data['precio'][indice]
        else:
            return None

    def get_quantity(self, data: list, indice: int)-> str:
        '''
        This method is used to get the quantity from the data
        '''
        if 'quantity' in data:
            return data['quantity'][indice]
        elif 'cantidad' in data:
            return data['cantidad'][indice]
        else:
            return None

    def get_created_by_crm_form(self, data: list, indice: int)-> str:
        '''
        This method is used to get the created by crm form from the data
        '''
        if 'created_by_crm_form' in data:
            return data['created_by_crm_form'][indice]
        elif 'creado_por_el_formulario_del_crm' in data:
            return data['creado_por_el_formulario_del_crm'][indice]
        else:
            return None

    def get_repeat_lead(self, data: list, indice: int)-> str:
        '''
        This method is used to get the repeat lead from the data
        '''
        if 'repeat_lead' in data:
            return False if data['repeat_lead'][indice] == 'N' else True
        elif 'prospecto_repetido' in data:
            return False if data['prospecto_repetido'][indice] == 'N' else True
        else:
            return False

    def get_client(self, data: list, indice: int)-> str:
        '''
        This method is used to get the client from the data
        '''
        if 'client' in data:
            return data['client'][indice]
        elif 'cliente' in data:
            return data['cliente'][indice]
        else:
            return None

    def get_customer_journey(self, data: list, indice: int)-> str:
        '''
        This method is used to get the customer journey from the data
        '''
        if 'customer_journey' in data:
            return data['customer_journey'][indice]
        elif 'recorrido_del_cliente' in data:
            return data['recorrido_del_cliente'][indice]
        else:
            return None

    def get_type(self, data: list, indice: int)-> str:
        '''
        This method is used to get the type from the data
        '''
        try:
            if 'type' in data:
                type = data['type'][indice]
            elif 'tipo' in data:
                type = data['tipo'][indice]
            else:
                return None

            return Type.objects.get(name=type)
        except Type.DoesNotExist:
            tipo = Type(name=type)
            tipo.save()
            return tipo
            # raise ValueError('The type does not exist: ' + type) from exc

    def get_country(self, data: list, indice: int)-> str:
        '''
        This method is used to get the country from the data
        '''
        try:
            if 'country' in data:
                country = data['country'][indice]
            elif 'pais' in data:
                country = data['country'][indice]
            else:
                return None

            return Country.objects.get(description=country)
        except Country.DoesNotExist:
            # raise ValueError('The country does not exist: ' + pais) from exc
            country = Country(description=country)
            country.save()
            return country

    def get_account(self, data: list, indice: int)-> str:
        '''
        This method is used to get the account from the data
        '''
        try:
            if 'account' in data:
                cuenta = data['account'][indice]
            elif 'cuenta' in data:
                cuenta = data['cuenta'][indice]
            else:
                return None

            return Account.objects.get(name=cuenta)
        except Account.DoesNotExist as exc:
            raise ValueError('The account does not exist: ' + cuenta) from exc

    def get_addl_type_details_other(self, data: list, indice: int)-> str:
        '''
        This method is used to get the addl_type_details_other from the data
        '''
        if 'addl_type_details_other' in data:
            return data['addl_type_details_other'][indice]
        else:
            return None

    def get_industry_sub_type(self, data: list, indice: int)-> str:
        '''
        This method is used to get the industry_sub_type from the data
        '''
        if 'industry_sub_type' in data:
            return data['industry_sub_type'][indice]
        else:
            return None

    def get_last_updated_on(self, data: list, indice: int)-> str:
        '''
        This method is used to get the last_updated_on from the data
        '''
        if 'last_updated_on' in data:
            return self.convertir_a_datetime(data['last_updated_on'][indice])
        elif 'última_actualización_en' in data:
            return self.convertir_a_datetime(data['última_actualización_en'][indice])
        else:
            return None

    def get_gg_sheet_row(self, data: list, indice: int)-> str:
        '''
        This method is used to get the gg_sheet_row from the data
        '''
        if 'gg_sheet_row' in data:
            return data['gg_sheet_row'][indice]
        else:
            return None

    def get_gg_sheet_id(self, data: list, indice: int)-> str:
        '''
        This method is used to get the gg_sheet_id from the data
        '''
        if 'gg_sheet_id' in data:
            return data['gg_sheet_id'][indice]
        else:
            return None

    def get_sub_type_hmcdelos(self, data: list, indice: int)-> str:
        '''
        This method is used to get the sub_type_hmcdelos from the data
        '''
        if 'sub_type_hmcdelos' in data:
            return data['sub_type_hmcdelos'][indice]
        else:
            return None

    def get_address(self, data: list, indice: int)-> str:
        """ This method is used to get the address from the data """
        if 'address' in data:
            return data['address'][indice]
        elif 'dirección' in data:
            return data['dirección'][indice]
        else:
            return None

    def get_street_house_no(self, data: list, indice: int)-> str:
        """ This method is used to get the street_house_no from the data """
        if 'street_house_no' in data:
            return data['street_house_no'][indice]
        elif 'calle_casa_núm' in data:
            return data['calle_casa_núm'][indice]
        else:
            return None

    def get_apartment_office_room_floor(self, data: list, indice: int)-> str:
        """ This method is used to get the apartment_office_room_floor from the data """
        if 'apartment_office_room_floor' in data:
            return data['apartment_office_room_floor'][indice]
        elif 'departamento_oficina_habitación_piso' in data:
            return data['departamento_oficina_habitación_piso'][indice]
        else:
            return None

    def get_city(self, data: list, indice: int)-> str:
        """ This method is used to get the city from the data """
        if 'city' in data:
            return data['city'][indice]
        elif 'ciudad' in data:
            return data['ciudad'][indice]
        else:
            return None

    def get_district(self, data: list, indice: int)-> str:
        """ This method is used to get the district from the data """
        if 'district' in data:
            return data['district'][indice]
        elif 'distrito' in data:
            return data['distrito'][indice]
        else:
            return None

    def get_regionarea(self, data: list, indice: int)-> str:
        """ This method is used to get the regionarea from the data """
        if 'regionarea' in data:
            return data['regionarea'][indice]
        elif 'regiónárea' in data:
            return data['regiónárea'][indice]
        else:
            return None

    def get_zippostal_code(self, data: list, indice: int)-> str:
        """ This method is used to get the zippostal_code from the data """
        if 'zippostal_code' in data:
            return data['zippostal_code'][indice]
        elif 'código_postal' in data:
            return data['código_postal'][indice]
        else:
            return None

    def get_country_dire(self, data: list, indice: int)-> str:
        """ This method is used to get the country_dire from the data """
        if 'country_dire' in data:
            return data['country_dire'][indice]
        elif 'país' in data:
            return data['país'][indice]
        else:
            return None
