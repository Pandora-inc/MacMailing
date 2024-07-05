""" This module contains the class for reading """
from datetime import datetime

import openpyxl
from django_select2.forms import Select2Widget

from auxiliares.models import ContactType, EmailType, SocialType, WebType, Country, Type
from .constants import INDICE_TRADUCCION, INDICE_TRADUCCION_CONTACT
from .models import (Clientes,
                     ClientesAddress,
                     ClientesContact,
                     ClientesEmail,
                     ClientesSocial,
                     ClientesWeb,
                     MailCorp,
                     Account)

class UtilExcelFile():
    '''
    Class for reading and manipulating Excel files.

    Methods:
    - __init__: Initializes the class with an empty workbook and active worksheet.
    - open_file: Opens an Excel file and sets the workbook and active worksheet.
    - clean_name: Cleans the name of a column by removing spaces, dashes, periods,
        commas, parentheses, and slashes.
    - get_structure: Returns a dictionary with the structure of the Excel file and a
        dictionary with the indices.
    - get_data: Returns a dictionary with the data of the Excel file and a dictionary
        with the indices.
    - print_datos: Prints the data of the Excel file to the console and creates and
        saves instances of various models in a Django project based on the data.
    - add_sheet: Adds a sheet to the Excel file.
    - add_row: Adds a row to the active worksheet.
    - save: Saves the Excel file.

    Fields:
    - file_name: The name of the Excel file.
    - wb: The openpyxl workbook object.
    - ws: The openpyxl worksheet object.
    '''
    file_name = None
    wb = None
    ws = None

    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def open_file(self, file_name):
        ''' Abre un archivo excel '''
        try:
            self.file_name = file_name
            self.wb = openpyxl.load_workbook(file_name)
            self.ws = self.wb.active
        except FileNotFoundError:
            print("El archivo no existe")
        except PermissionError:
            print("No tiene permisos para abrir el archivo")
        except IsADirectoryError:
            print("El archivo es un directorio")

    def clean_name(self, name):
        ''' Limpia el nombre de los campos '''
        name = str(name).lower()
        name = name.replace(" ", "_")
        name = name.replace("-", "_")
        name = name.replace(".", "")
        name = name.replace(",", "")
        name = name.replace("(", "")
        name = name.replace(")", "")
        name = name.replace("/", "")
        return name

    def traducir_claves_dict(self, clave, diccionario):
        '''
        Diccionario de traducción
        '''
        if clave in diccionario:
            return diccionario[clave]
        return clave

    def convertir_a_datetime(self, fecha_str):
        """ Convierte una fecha en string a datetime """
        formatos = ["%Y-%m-%d",
                    "%d/%m/%Y",
                    "%m-%d-%Y",
                    "%m/%d/%y %H:%M",
                    "%m/%d/%y %H:%M:%S %p",
                    "%m/%d/%Y %H:%M",
                    "%m/%d/%Y %H:%M:%S %p"]
        for formato in formatos:
            try:
                if fecha_str is None:
                    return None
                if isinstance(fecha_str, datetime):
                    return fecha_str
                fecha_datetime = datetime.strptime(fecha_str, formato)
                return fecha_datetime
            except ValueError:
                print(f"No se pudo convertir la fecha {fecha_str} con el formato {formato}")

        raise ValueError(f"No se pudo encontrar un formato válido para la fecha: {fecha_str}")



    def get_structure(self):
        ''' Devuelve un diccionario con la estructura del excel y un diccionario con los indices '''
        response = {}
        indice = {}
        for row in self.ws.iter_rows(min_row=1, max_row=1):
            i = 0
            for cell in row:
                indice[i] = cell.value

                if cell.value:
                    indio = self.clean_name(cell.value)
                    indio = self.traducir_claves_dict(indio, INDICE_TRADUCCION)
                    # TODO: Este if es una excepción, hay que ver como solucionarlo
                    if i == 58:
                        indio = indio+"_dire"
                else:
                    indio = i

                response[indio] = []
                i += 1

        return response, indice

    def get_data(self) -> tuple:
        ''' Devuelve un diccionario con los datos del excel y un diccionario con los indices '''
        try:
            estructura, indice = self.get_structure()

            for row in self.ws.iter_rows(min_row=2):
                i = 0
                for campo in row:
                    if indice[i]:
                        indio = self.clean_name(indice[i])
                        indio = self.traducir_claves_dict(indio, INDICE_TRADUCCION)
                        indio = indio + "_dire" if i == 58 else indio
                    else:
                        indio = i
                    dato = campo.value
                    estructura[indio].append(dato)
                    i += 1

        except KeyError as e_error:
            print("Error de key error: " + str(e_error))
            raise ValueError(str(e_error)) from e_error
        except Exception as e_error:
            print("Error al obtener los datos: " + str(e_error))
            raise ValueError(str(e_error)) from e_error

        return estructura, indice

    def print_datos(self):
        ''' Imprime los datos del excel '''
        data, indice = self.get_data()

        for i in range(len(data[list(data.keys())[0]])):
            if data['id'][i] is None:
                print("No hay id en posición: " + str(i))
                continue

            cliente = self.get_or_create_cliente(data['id'][i])

            # Asignar atributos al cliente
            self.assign_attributes_to_cliente(cliente, data, i)

            # Guardar el cliente
            cliente.save()

            # Asignar dirección al cliente
            self.assign_direccion_to_cliente(cliente, data, i)

            # Guardar la dirección
            cliente.clientesaddress.save()

            # Asignar contactos, webs, emails, sociales al cliente
            self.assign_contacts_to_cliente(cliente, data, indice, i)

            print(f"Se ha {'creado' if cliente.pk is None else 'actualizado'} \
                  el cliente: {cliente.first_name} {cliente.last_name}")

    def get_or_create_cliente(self, cliente_id):
        """ Obtiene o crea un cliente """
        try:
            return Clientes.objects.get(cliente_id=cliente_id)
        except Clientes.DoesNotExist:
            return Clientes(cliente_id=cliente_id)

    def assign_attributes_to_cliente(self, cliente, data, i):
        """ Asigna los atributos al cliente """
        cliente.status = self.get_status(data, i)
        cliente.lead_name = self.get_lead_name(data, i)
        cliente.salutation = self.get_salutation(data, i)
        cliente.first_name = self.get_first_name(data, i)
        cliente.middle_name = self.get_middle_name(data, i)
        cliente.last_name = self.get_last_name(data, i)
        cliente.date_of_birth = self.get_date_of_birth(data, i)
        cliente.created = self.get_created(data, i)
        cliente.source = self.get_source(data, i)
        cliente.responsible = self.get_responsible(data, i)
        cliente.status_information = self.get_status_information(data, i)
        cliente.source_information = self.get_source_information(data, i)
        cliente.created_by = self.get_created_by(data, i)
        cliente.modified = self.get_modified(data, i)
        cliente.modified_by = self.get_modified_by(data, i)
        cliente.company_name = self.get_company_name(data, i)
        cliente.position = self.get_position(data, i)
        cliente.comment = self.get_comment(data, i)
        cliente.total = self.get_total(data, i)
        cliente.currency = self.get_currency(data, i)
        cliente.product = self.get_product(data, i)
        cliente.price = self.get_price(data, i)
        cliente.quantity = self.get_quantity(data, i)
        cliente.created_by_crm_form = self.get_created_by_crm_form(data, i)
        cliente.repeat_lead = self.get_repeat_lead(data, i)
        cliente.client = self.get_client(data, i)
        cliente.customer_journey = self.get_customer_journey(data, i)
        cliente.type = self.get_type(data, i)
        cliente.country = self.get_country(data, i)
        cliente.account = self.get_account(data, i)
        cliente.addl_type_details_other = self.get_addl_type_details_other(data, i)
        cliente.industry_sub_type = self.get_industry_sub_type(data, i)
        cliente.last_updated_on = self.get_last_updated_on(data, i)

    def assign_direccion_to_cliente(self, cliente, data, i):
        """ Asigna la dirección al cliente """
        direccion = ClientesAddress(
            cliente=cliente,
            address=self.get_address(data, i),
            street_house_no=self.get_street_house_no(data, i),
            apartment_office_room_floor=self.get_apartment_office_room_floor(data, i),
            city=self.get_city(data, i),
            district=self.get_district(data, i),
            region_area=self.get_regionarea(data, i),
            postal_code=self.get_zippostal_code(data, i),
            country=self.get_country_dire(data, i)
        )
        cliente.clientesaddress = direccion

    def assign_contacts_to_cliente(self, cliente, data, indice, i):
        """ Asigna los contactos al cliente """
        for e, (_, value) in enumerate(data.items()):
            if value[i]:
                nombre = self.traducir_claves_dict(indice[e], INDICE_TRADUCCION_CONTACT)
                if nombre:
                    self.add_contact_to_cliente(cliente, nombre, value[i])

    def add_contact_to_cliente(self, cliente: Clientes, nombre: str, valor: str):
        """ Agrega un contacto al cliente """
        if ContactType.objects.filter(name=nombre).exists():
            contact_type = ContactType.objects.get(name=nombre)
            if not ClientesContact.objects.filter(cliente=cliente, type=contact_type).exists():
                cliente.add_contact(contact_type, valor)
            else:
                contact = ClientesContact.objects.get(cliente=cliente, type=contact_type)
                contact.data = valor
                contact.save()
        elif WebType.objects.filter(name=nombre).exists():
            web_type = WebType.objects.get(name=nombre)
            if not ClientesWeb.objects.filter(cliente=cliente, type=web_type).exists():
                cliente.add_web(web_type, valor)
            else:
                web = ClientesWeb.objects.get(cliente=cliente, type=web_type)
                web.data = valor
                web.save()
        elif EmailType.objects.filter(name=nombre).exists():
            email_type = EmailType.objects.get(name=nombre)
            if not ClientesEmail.objects.filter(cliente=cliente, type=email_type).exists():
                cliente.add_email(email_type, valor)
            else:
                email = ClientesEmail.objects.get(cliente=cliente, type=email_type)
                email.data = valor
                email.save()
        elif SocialType.objects.filter(name=nombre).exists():
            social_type = SocialType.objects.get(name=nombre)
            if not ClientesSocial.objects.filter(cliente=cliente, type=social_type).exists():
                cliente.add_social(social_type, valor)
            else:
                social = ClientesSocial.objects.get(cliente=cliente, type=social_type)
                social.data = valor
                social.save()


    def add_sheet(self, sheet_name):
        ''' Agrega una hoja al archivo'''
        self.ws = self.wb.create_sheet(sheet_name)

    def add_row(self, row):
        ''' Agrega una fila al archivo '''
        self.ws.append(row)

    def save(self):
        ''' Guarda el archivo '''
        self.wb.save(self.file_name)

    def get_status(self, data: list, indice: int)-> str:
        '''
        This method is used to get the status from the data
        '''
        if 'status' in data:
            return data['status'][indice]
        if 'estatus' in data:
            return data['estatus'][indice]
        if 'stage' in data:
            return data['stage'][indice]
        if 'etapa' in data:
            return data['etapa'][indice]
        return None

    def get_lead_name(self, data: list, indice: int)-> str:
        '''
        This method is used to get the lead name from the data
        '''
        if 'lead_name' in data:
            return data['lead_name'][indice]
        if 'título_de_prospecto' in data:
            return data['título_de_prospecto'][indice]
        return None

    def get_salutation(self, data: list, indice: int)-> str:
        '''
        This method is used to get the salutation from the data
        '''
        if 'salutation' in data:
            return data['salutation'][indice]
        if 'saludo' in data:
            return data['saludo'][indice]
        return None

    def get_first_name(self, data: list, indice: int)-> str:
        '''
        This method is used to get the first name from the data
        '''
        if 'first_name' in data:
            return data['first_name'][indice]
        if 'nombre' in data:
            return data['nombre'][indice]
        return None

    def get_middle_name(self, data: list, indice: int)-> str:
        '''
        This method is used to get the middle name from the data
        '''
        if 'middle_name' in data:
            return data['middle_name'][indice]
        if 'segundo_nombre' in data:
            return data['segundo_nombre'][indice]
        return None

    def get_last_name(self, data: list, indice: int)-> str:
        '''
        This method is used to get the last name from the data
        '''
        if 'last_name' in data:
            return data['last_name'][indice]
        if 'apellido' in data:
            return data['apellido'][indice]
        return None

    def get_date_of_birth(self, data: list, indice: int)-> str:
        '''
        This method is used to get the date of birth from the data
        '''
        if 'date_of_birth' in data:
            return data['date_of_birth'][indice]
        if 'fecha_de_nacimiento' in data:
            return data['fecha_de_nacimiento'][indice]
        return None

    def get_created(self, data: list, indice: int)-> str:
        '''
        This method is used to get the created from the data
        '''
        if 'created' in data:
            return self.convertir_a_datetime(data['created'][indice])
        if 'creado' in data:
            return self.convertir_a_datetime(data['creado'][indice])
        return None

    def get_source(self, data: list, indice: int)-> str:
        '''
        This method is used to get the source from the data
        '''
        if 'source' in data:
            return data['source'][indice]
        if 'origen' in data:
            return data['origen'][indice]
        return None

    def get_responsible(self, data: list, indice: int)-> str:
        '''
        This method is used to get the responsible from the data
        '''
        try:
            if 'responsible' in data:
                responsible = data['responsible'][indice]
            elif 'responsable' in data:
                responsible = data['responsable'][indice]
            else:
                return None

            return MailCorp.objects.get(name=responsible)
        except MailCorp.DoesNotExist as exc:
            raise ValueError('The responsible does not exist: ' + responsible) from exc

    def get_status_information(self, data: list, indice: int)-> str:
        '''
        This method is used to get the status information from the data
        '''
        if 'status_information' in data:
            return data['status_information'][indice]
        if 'información_de_estado' in data:
            return data['información_de_estado'][indice]
        return None

    def get_source_information(self, data: list, indice: int)-> str:
        '''
        This method is used to get the source information from the data
        '''
        try:
            if 'source_information' in data:
                return data['source_information'][indice]
            if 'información_de_origen' in data:
                return data['información_de_origen'][indice]
            return None
        except MailCorp.DoesNotExist as exc:
            raise ValueError('The source_information does not exist: ' +
                             data['información_de_origen'][indice]) from exc
        except Exception as exc:
            raise ValueError('Error en source_information: ' +
                             data['información_de_origen'][indice]) from exc

    def get_created_by(self, data: list, indice: int)-> str:
        '''
        This method is used to get the created by from the data
        '''
        try:
            if 'created_by' in data:
                created_by = data['created_by'][indice]
            elif 'creado_por' in data:
                created_by = data['creado_por'][indice]
            else:
                return None

            return MailCorp.objects.get(name=created_by)
        except MailCorp.DoesNotExist as exc:
            raise ValueError('The created_by does not exist: ' + created_by) from exc

    def get_modified(self, data: list, indice: int)-> str:
        '''
        This method is used to get the modified from the data
        '''
        if 'modified' in data:
            return self.convertir_a_datetime(data['modified'][indice])
        if 'modificado' in data:
            return self.convertir_a_datetime(data['modificado'][indice])
        return None

    def get_modified_by(self, data: list, indice: int)-> str:
        '''
        This method is used to get the modified by from the data
        '''
        try:
            if 'modified_by' in data:
                modified_by = data['modified_by'][indice]
            elif 'modificado_por' in data:
                modified_by = data['modificado_por'][indice]
            else:
                return None

            return MailCorp.objects.get(name=modified_by)
        except MailCorp.DoesNotExist as exc:
            raise ValueError('The modified_by does not exist: ' + modified_by) from exc

    def get_company_name(self, data: list, indice: int)-> str:
        '''
        This method is used to get the company name from the data
        '''
        if 'company_name' in data:
            return data['company_name'][indice]
        if 'nombre_de_la_compañía' in data:
            return data['nombre_de_la_compañía'][indice]
        return None

    def get_position(self, data: list, indice: int)-> str:
        '''
        This method is used to get the position from the data
        '''
        if 'position' in data:
            return data['position'][indice]
        if 'cargo' in data:
            return data['cargo'][indice]
        return None

    def get_comment(self, data: list, indice: int)-> str:
        '''
        This method is used to get the comment from the data
        '''
        if 'comment' in data:
            return data['comment'][indice]
        if 'comentario' in data:
            return data['comentario'][indice]
        return None

    def get_total(self, data: list, indice: int)-> str:
        '''
        This method is used to get the total from the data
        '''
        if 'total' in data:
            return data['total'][indice]
        return None

    def get_currency(self, data: list, indice: int)-> str:
        '''
        This method is used to get the currency from the data
        '''
        if 'currency' in data:
            return data['currency'][indice]
        if 'moneda' in data:
            return data['moneda'][indice]
        return None

    def get_product(self, data: list, indice: int)-> str:
        '''
        This method is used to get the product from the data
        '''
        if 'product' in data:
            return data['product'][indice]
        if 'producto' in data:
            return data['producto'][indice]
        return None

    def get_price(self, data: list, indice: int)-> str:
        '''
        This method is used to get the price from the data
        '''
        if 'price' in data:
            return data['price'][indice]
        if 'precio' in data:
            return data['precio'][indice]
        return None

    def get_quantity(self, data: list, indice: int)-> str:
        '''
        This method is used to get the quantity from the data
        '''
        if 'quantity' in data:
            return data['quantity'][indice]
        if 'cantidad' in data:
            return data['cantidad'][indice]
        return None

    def get_created_by_crm_form(self, data: list, indice: int)-> str:
        '''
        This method is used to get the created by crm form from the data
        '''
        if 'created_by_crm_form' in data:
            return data['created_by_crm_form'][indice]
        if 'creado_por_el_formulario_del_crm' in data:
            return data['creado_por_el_formulario_del_crm'][indice]
        return None

    def get_repeat_lead(self, data: list, indice: int) -> str:
        '''
        Este método se utiliza para obtener el "lead" repetido de los datos
        '''
        if 'repeat_lead' in data:
            return data['repeat_lead'][indice] != 'N'
        if 'prospecto_repetido' in data:
            return data['prospecto_repetido'][indice] != 'N'
        return False

    def get_client(self, data: list, indice: int)-> str:
        '''
        This method is used to get the client from the data
        '''
        if 'client' in data:
            return data['client'][indice]
        if 'cliente' in data:
            return data['cliente'][indice]
        return None

    def get_customer_journey(self, data: list, indice: int)-> str:
        '''
        This method is used to get the customer journey from the data
        '''
        if 'customer_journey' in data:
            return data['customer_journey'][indice]
        if 'recorrido_del_cliente' in data:
            return data['recorrido_del_cliente'][indice]
        return None

    def get_type(self, data: list, indice: int)-> str:
        '''
        This method is used to get the type from the data
        '''
        try:
            if 'type' in data:
                tipo = data['type'][indice]
            elif 'tipo' in data:
                tipo = data['tipo'][indice]
            else:
                return None

            return Type.objects.get(name=tipo)
        except Type.DoesNotExist:
            tipo = Type(name=tipo)
            tipo.save()
            return tipo
            # raise ValueError('The type does not exist: ' + type) from exc

    def get_country(self, data: list, indice: int)-> str:
        '''
        This method is used to get the country from the data
        '''
        try:
            if 'country' in data:
                country = data['country'][indice]
            elif 'pais' in data:
                country = data['country'][indice]
            else:
                return None

            return Country.objects.get(description=country)
        except Country.DoesNotExist:
            # raise ValueError('The country does not exist: ' + pais) from exc
            country = Country(description=country)
            country.save()
            return country

    def get_account(self, data: list, indice: int)-> str:
        '''
        This method is used to get the account from the data
        '''
        try:
            if 'account' in data:
                cuenta = data['account'][indice]
            elif 'cuenta' in data:
                cuenta = data['cuenta'][indice]
            else:
                return None

            return Account.objects.get(name=cuenta)
        except Account.DoesNotExist as exc:
            raise ValueError('The account does not exist: ' + cuenta) from exc

    def get_addl_type_details_other(self, data: list, indice: int)-> str:
        '''
        This method is used to get the addl_type_details_other from the data
        '''
        if 'addl_type_details_other' in data:
            return data['addl_type_details_other'][indice]
        return None

    def get_industry_sub_type(self, data: list, indice: int)-> str:
        '''
        This method is used to get the industry_sub_type from the data
        '''
        if 'industry_sub_type' in data:
            return data['industry_sub_type'][indice]
        return None

    def get_last_updated_on(self, data: list, indice: int)-> str:
        '''
        This method is used to get the last_updated_on from the data
        '''
        if 'last_updated_on' in data:
            return self.convertir_a_datetime(data['last_updated_on'][indice])
        if 'última_actualización_en' in data:
            return self.convertir_a_datetime(data['última_actualización_en'][indice])
        return None

    def get_gg_sheet_row(self, data: list, indice: int)-> str:
        '''
        This method is used to get the gg_sheet_row from the data
        '''
        if 'gg_sheet_row' in data:
            return data['gg_sheet_row'][indice]
        return None

    def get_gg_sheet_id(self, data: list, indice: int)-> str:
        '''
        This method is used to get the gg_sheet_id from the data
        '''
        if 'gg_sheet_id' in data:
            return data['gg_sheet_id'][indice]
        return None

    def get_sub_type_hmcdelos(self, data: list, indice: int)-> str:
        '''
        This method is used to get the sub_type_hmcdelos from the data
        '''
        if 'sub_type_hmcdelos' in data:
            return data['sub_type_hmcdelos'][indice]
        return None

    def get_address(self, data: list, indice: int)-> str:
        """ This method is used to get the address from the data """
        if 'address' in data:
            return data['address'][indice]
        if 'dirección' in data:
            return data['dirección'][indice]
        return None

    def get_street_house_no(self, data: list, indice: int)-> str:
        """ This method is used to get the street_house_no from the data """
        if 'street_house_no' in data:
            return data['street_house_no'][indice]
        if 'calle_casa_núm' in data:
            return data['calle_casa_núm'][indice]
        return None

    def get_apartment_office_room_floor(self, data: list, indice: int)-> str:
        """ This method is used to get the apartment_office_room_floor from the data """
        if 'apartment_office_room_floor' in data:
            return data['apartment_office_room_floor'][indice]
        if 'departamento_oficina_habitación_piso' in data:
            return data['departamento_oficina_habitación_piso'][indice]
        return None

    def get_city(self, data: list, indice: int)-> str:
        """ This method is used to get the city from the data """
        if 'city' in data:
            return data['city'][indice]
        if 'ciudad' in data:
            return data['ciudad'][indice]
        return None

    def get_district(self, data: list, indice: int)-> str:
        """ This method is used to get the district from the data """
        if 'district' in data:
            return data['district'][indice]
        if 'distrito' in data:
            return data['distrito'][indice]
        return None

    def get_regionarea(self, data: list, indice: int)-> str:
        """ This method is used to get the regionarea from the data """
        if 'regionarea' in data:
            return data['regionarea'][indice]
        if 'regiónárea' in data:
            return data['regiónárea'][indice]
        return None

    def get_zippostal_code(self, data: list, indice: int)-> str:
        """ This method is used to get the zippostal_code from the data """
        if 'zippostal_code' in data:
            return data['zippostal_code'][indice]
        if 'código_postal' in data:
            return data['código_postal'][indice]
        return None

    def get_country_dire(self, data: list, indice: int)-> str:
        """ This method is used to get the country_dire from the data """
        if 'country_dire' in data:
            return data['country_dire'][indice]
        if 'país' in data:
            return data['país'][indice]
        return None



def if_admin(user):
    """ Función para verificar si el usuario es admin """
    if user.is_superuser:
        return True

    grupos = user.groups.all()

    for grupo in grupos:
        if grupo.name in ('SuperAdmin', 'Admin'):
            return True
    return False


def get_response_account(user):
    ''' Obtener las cuentas de respuesta del usuario '''
    return MailCorp.objects.filter(user=user)


class Select2WidgetWithSearch(Select2Widget):
    """ Select2Widget with search fields """
    search_fields = [
        'text__icontains',
    ]
